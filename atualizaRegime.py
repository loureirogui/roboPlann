import traceback
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from docx import Document
import openpyxl
import unicodedata
import re
from art import *



# Exibir texto simples em ASCII
tprint("Robô Regime")


# Solicita as credenciais do usuário
nomeCliente = input('Qual o nome do escritório? ;')
emailLogin = input('Qual o email do Acessorias? ;')
senhaLogin = input('Qual a senha do Acessorias? ;')
nomePlanilha = input('Qual o nome da planilha do clente? ;')
print("Acessando Acessorias Agora")


erro_log = f"erros_{nomeCliente}_regime.txt"

def registrar_erro(mensagem):
    with open(erro_log, "a", encoding="utf-8") as file:
        file.write(mensagem + "\n")


# Caminho para o driver do Edge
driver_path = 'msedgedriver.exe'

# Configura as opções do Edge
edge_options = Options()
edge_options.headless = False  # Executa o Edge em modo não headless

# Inicializa o navegador Edge
service = Service(driver_path)
edge_driver = webdriver.Edge(service=service, options=edge_options)

# Abre o link desejado
url = f"https://app.acessorias.com/sysmain.php?m=22"
edge_driver.get(url)

# Lógica de login
try:
    # Espera o campo de e-mail aparecer
    email_input = WebDriverWait(edge_driver, 10).until(
        EC.visibility_of_element_located((By.NAME, 'mailAC'))
    )
    # Insere o e-mail no campo
    email_input.send_keys(emailLogin)
except:
    print("Erro ao inserir o e-mail no campo de login:")
    registrar_erro("Erro ao inserir o e-mail no campo de login:")

try:
    # Espera o campo de senha aparecer
    senha_input = WebDriverWait(edge_driver, 10).until(
        EC.visibility_of_element_located((By.NAME, 'passAC'))
    )
    # Insere a senha no campo
    senha_input.send_keys(senhaLogin)
except:
    print("Erro ao inserir a senha no campo de senha:")
    registrar_erro("Erro ao inserir a senha no campo de login:")


try:
    # Espera o botão de login aparecer e ser clicável
    login_button = WebDriverWait(edge_driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="site-corpo"]/section[1]/div/form/div[2]/button'))
    )
    # Clique no botão de login
    login_button.click()
    registrar_erro("Erro ao clicar no botão de login:")
except:
    print("Erro ao clicar no botão de login:")
    registrar_erro("Erro ao clicar no botão de login:")

time.sleep(3)

url = 'https://app.acessorias.com/sysmain.php?m=22'
edge_driver.get(url)

# Carrega o arquivo .xlsx
workbook = openpyxl.load_workbook(nomePlanilha)

# Seleciona a planilha
sheet = workbook['Regime tributário']

# INATIVADOR DE REGIMES PRÉ-DEFINIDOS
try:
     # Espera que os regimes tributários estejam visíveis
    regimesPadrao = WebDriverWait(edge_driver, 10).until(
        EC.visibility_of_all_elements_located((By.CSS_SELECTOR, '.dRow, .dOdd'))
    )

    for regimePadrao in regimesPadrao:
        try:
            edge_driver.get(url)
            regime_links = WebDriverWait(edge_driver, 10).until(
                EC.visibility_of_all_elements_located((By.XPATH, '//*[@id="main-container"]/div[2]/div[2]/div/div/div[4]/div[1]/a'))
            )
            for regime in regime_links:
                regime.click()  # Clica no regime tributário 
                try:
                    # Espera o campo de inativação aparecer
                    inativadorObr = WebDriverWait(edge_driver, 10).until(
                        EC.visibility_of_element_located((By.NAME, 'RegAtivo'))
                    )
                    # Seleciona a opção "Não" no campo de inativação
                    select = Select(inativadorObr)
                    select.select_by_visible_text('Não')
                    
                except Exception as e:
                    print("Erro ao selecionar o não")
                    registrar_erro("Erro ao selecionar o não para inativar regime tributário")
                
                # Clica no botão de salvar
                try:
                    # Espera o campo de inativação aparecer
                    saveButton = WebDriverWait(edge_driver, 10).until(
                        EC.visibility_of_element_located((By.XPATH, '//*[@id="main-container"]/div[2]/div[2]/div/div/form/div[1]/div[3]/button[1]'))
                    )
                    saveButton.click()  # Descomentar para salvar de fato
                    time.sleep(2)
                    print("Regime tributário inativado com sucesso.")
                    registrar_erro("Regime tributário inativado com sucesso.")
                except Exception as e:
                    print("Erro ao clicar no botão de salvar o inativador:")
                    registrar_erro("Erro ao clicar no botão de salvar o inativador:")
        except Exception as e:
            print("Erro ao clicar nos regimes tributários:")
            registrar_erro("Erro ao clicar nos regimes tributários:")
except Exception as e:
    print(" ")

edge_driver.get(url)

# Criar novos regimes tributários...  
# após localizar os regimes na planilha,                      
# Itera sobre as colunas da primeira linha até encontrar uma célula vazia

 # Função para normalizar o texto
def normalize_text(text):
    if text is None:
        return ""
    # Remove acentos e caracteres diacríticos
    text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    # Substitui o símbolo ordinal "º" e "°" por "o"
    text = text.replace('º', '').replace('°', '').replace('ª', '')
    # Remove espaços extras e coloca o texto em minúsculas
    text = text.strip().lower()
    return text

col = 1
while True:
    titleRegime = sheet.cell(row=2, column=col).value
    if titleRegime is None:  # Para quando encontrar uma célula vazia
        break

    titleNormalizado = normalize_text(titleRegime)

    # Navegar para a URL antes de tentar adicionar novos regimes
    url = 'https://app.acessorias.com/sysmain.php?m=22'
    edge_driver.get(url)

    # Processa cada regime
    while titleRegime:
        try:
            url = 'https://app.acessorias.com/sysmain.php?m=22'
            edge_driver.get(url)
            titleRegime = sheet.cell(row=2, column=col).value
            # Encontra e clica no botão de criar novo regime tributário
            criarRegimeButton = WebDriverWait(edge_driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="main-container"]/div[2]/div[2]/div/div/div[1]/button'))
            )
            criarRegimeButton.click()
            print("Cliquei no botão para criar um novo regime com sucesso.")

            # Encontra e preenche o nome do regime tributário
            nomeRegime = WebDriverWait(edge_driver, 10).until(
                EC.visibility_of_element_located((By.NAME, 'RegNome'))
            )
            nomeRegime.send_keys(titleRegime)
            print(titleRegime)
            print("Nome do regime inserido com sucesso.")

            # Clica no botão de salvar
            saveButton = WebDriverWait(edge_driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="main-container"]/div[2]/div[2]/div/div/form/div/div[3]/button[1]'))
            )
            saveButton.click()  # Descomentar para salvar de fato
            time.sleep(2)
            print(f"Regime tributário;{titleRegime}; criado com sucesso.")
            registrar_erro(f"Regime tributário;{titleRegime}; criado com sucesso.")
        except Exception as e:
            print(f"Erro ao processar o regime tributário; {titleRegime}")
            registrar_erro(f"Erro ao processar o regime tributário; {titleRegime}")

        # URL para adicionar obrigações
        url = 'https://app.acessorias.com/sysmain.php?m=23&act=e&tr=R&i=42&p=0&o=RegNome&rpp=14'
        edge_driver.get(url)

        row = 3  # Iniciando na linha 3
        empty_cell_count = 0  # Contador de células vazias consecutivas

        while True:
            obrigacaoName = sheet.cell(row=row, column=col).value

            if obrigacaoName is None:
                empty_cell_count += 1
                if empty_cell_count >= 5:
                    break
            else:
                empty_cell_count = 0  # Reseta o contador se encontrar uma célula não vazia
                obrigacaoNameNormalizado = normalize_text(obrigacaoName)

                try:
                    # Espera o campo de inativação aparecer
                    seletorObr = WebDriverWait(edge_driver, 10).until(
                        EC.visibility_of_element_located((By.XPATH, '//*[@id="newObr"]'))
                    )
                    # Seleciona a opção com nome da obrigação no campo de inativação
                    select = Select(seletorObr)
                    options = select.options
                    for option in options:
                        optionNormalizado = normalize_text(option.text)
                        if optionNormalizado.startswith(obrigacaoNameNormalizado):
                            select.select_by_visible_text(option.text)
                            time.sleep(1)
                            break

                    # Clica no botão de adicionar
                    addButton = WebDriverWait(edge_driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="divSelectObr"]/button'))
                    )
                    actions = ActionChains(edge_driver)
                    actions.move_to_element(addButton).click().perform()
                    print(f"Obrigação; {obrigacaoName}; adicionada com sucesso.")
                    registrar_erro(f"Obrigação; {obrigacaoName}; adicionada com sucesso no regmime; {titleRegime}")
                except Exception as e:
                    print(f"Erro ao selecionar a obrigação {obrigacaoName}")
                    registrar_erro(f"Erro ao adicionar; {obrigacaoName}; no regmime; {titleRegime}")
                    
                # Clica no botão de salvar
            row += 1  # Vai para a próxima linha
        saveButton = WebDriverWait(edge_driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="main-container"]/div[2]/div[2]/div/div/form/div[1]/div[3]/button[1]'))
        )
        saveButton.click()  # Descomentar para salvar de fato
        time.sleep(2)
        print(f"Regime tributário salvo com sucesso; {titleRegime}")
        registrar_erro(f"Regime tributário salvo com sucesso; {titleRegime}")
                    
    # Move para a próxima coluna
        col += 1